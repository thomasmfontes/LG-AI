import pandas as pd
from openpyxl import load_workbook
from functools import lru_cache
from typing import Dict, List
from pathlib import Path

from config import Config
from .utils import validate_file_exists, normalize_campo, InvalidDataError
from .logger import setup_logger

logger = setup_logger(__name__)


class DataLoader:
    """Carrega e gerencia dados das planilhas Excel"""
    
    def __init__(self):
        self.df_redes: pd.DataFrame = None
        self.df_campos: pd.DataFrame = None
        self.comentarios: Dict[str, str] = {}
        self.mapa_rede_canal: Dict[str, str] = {}
        self._loaded = False
    
    def load_all(self) -> None:
        """Carrega todos os dados necessários"""
        if self._loaded:
            logger.info("Dados já carregados, usando cache")
            return
        
        logger.info("Iniciando carregamento de dados...")
        
        try:
            self._validate_files()
            self.df_redes = self._load_redes()
            self.df_campos = self._load_and_normalize_campos()
            self.comentarios = self._load_comentarios()
            self.mapa_rede_canal = self._create_rede_canal_map()
            
            self._loaded = True
            logger.info("Dados carregados com sucesso")
            
        except Exception as e:
            logger.error(f"Erro ao carregar dados: {e}")
            raise
    
    def _validate_files(self) -> None:
        """Valida existência de todos os arquivos necessários"""
        files_to_check = [
            Config.REDES_FILE,
            Config.CAMPOS_FILE,
            Config.MODELO_FILE,
            Config.MANUAL_FILE
        ]
        
        for filepath in files_to_check:
            validate_file_exists(filepath)
            logger.debug(f"Arquivo validado: {filepath.name}")
    
    def _load_redes(self) -> pd.DataFrame:
        """Carrega planilha de redes"""
        logger.info(f"Carregando redes de {Config.REDES_FILE.name}")
        df = pd.read_excel(Config.REDES_FILE)
        
        # Validação
        required_columns = ['Rede', 'Canal']
        missing = set(required_columns) - set(df.columns)
        if missing:
            raise InvalidDataError(f"Colunas faltando em Redes: {missing}")
        
        if df['Rede'].isna().any():
            raise InvalidDataError("Existem redes com valores nulos")
        
        logger.info(f"Carregadas {len(df)} redes")
        return df
    
    def _load_and_normalize_campos(self) -> pd.DataFrame:
        """Carrega e normaliza planilha de campos"""
        logger.info(f"Carregando campos de {Config.CAMPOS_FILE.name}")
        df = pd.read_excel(Config.CAMPOS_FILE)
        
        # Validação
        if 'CAMPO' not in df.columns:
            raise InvalidDataError("Coluna 'CAMPO' não encontrada")
        
        if df['CAMPO'].isna().any():
            raise InvalidDataError("Existem campos com valores nulos")
        
        # Normalização
        df['CAMPO_NORMALIZADO'] = df['CAMPO'].apply(normalize_campo)
        
        logger.info(f"Carregados {len(df)} campos")
        logger.debug(f"Campos normalizados: {df['CAMPO_NORMALIZADO'].tolist()[:5]}...")
        
        return df
    
    def _load_comentarios(self) -> Dict[str, str]:
        """Extrai comentários do modelo Excel"""
        logger.info(f"Extraindo comentários de {Config.MODELO_FILE.name}")
        comentarios = {}
        
        wb = load_workbook(Config.MODELO_FILE, data_only=True)
        ws = wb.active
        
        for cell in ws[1]:
            if cell.comment and cell.value:
                key = cell.value.strip().lower()
                comentarios[key] = cell.comment.text.strip()
        
        logger.info(f"Extraídos {len(comentarios)} comentários")
        return comentarios
    
    def _create_rede_canal_map(self) -> Dict[str, str]:
        """Cria mapeamento de rede para canal"""
        mapa = dict(zip(self.df_redes['Rede'], self.df_redes['Canal']))
        logger.debug(f"Criado mapa com {len(mapa)} redes")
        return mapa
    
    def get_lista_redes(self) -> List[str]:
        """Retorna lista ordenada de redes"""
        return sorted(self.mapa_rede_canal.keys())
    
    def get_lista_campos(self) -> List[str]:
        """Retorna lista de campos normalizados em uppercase"""
        return self.df_campos['CAMPO_NORMALIZADO'].str.upper().tolist()
    
    @lru_cache(maxsize=128)
    def get_canal_for_rede(self, rede: str) -> str:
        """
        Obtém canal para uma rede (com cache).
        
        Args:
            rede: Nome da rede
        
        Returns:
            Nome do canal
        """
        canal_original = self.mapa_rede_canal.get(rede, "").strip().upper()
        return Config.MAPEAMENTO_CANAIS.get(canal_original, canal_original)
